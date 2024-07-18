from datetime import datetime
import pandas as pd
import os
import openpyxl 
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
import openpyxl.utils.dataframe as op    # Module needed for creating new workbook
#from openpyxl import load_workbook   # Module needed for loading existing workbook
import configparser

config = configparser.ConfigParser()
config.read('Data//config.ini')

def writeLog(log, name="", descriptor=".txt"):
    now = datetime.now()
    dt_string = now.strftime(" %d-%m-%Y %H %M")
    with open(f"Logs/{name}{dt_string}{descriptor}", 'w') as file:
        for row in log:
            file.write(row)
            file.write("\n")

def logModel(modelType, model, train_stats, test_stats, data, params, fieldModelType):
    # weights = [w for w in model.w]
    
    # log.append("Weights:")

    # log += [",".join([str(w) for w in class_w]) for class_w in weights]

    log = list()
    log.append(f"Model Type: {modelType}")
    log.append("")
    log.append(f"Training Size = {len(data[0])}")
    log.append(f"Testing Size = {len(data[2])}")
    log.append("")
    log.append(f"Training Accuracy = {train_stats[0]}")
    log.append(f"Testing Accuracy = {test_stats[0]}")
    log.append("")
    log.append(f"Training Average Error = {train_stats[1]}")
    log.append(f"Testing Average Error = {test_stats[1]}")
    log.append("")
    log.append(f"Training Recall = {train_stats[2]}")
    log.append(f"Testing Recall = {test_stats[2]}")
    log.append("")
    log.append(f"Training f1 (micro, macro, weighted) = {train_stats[3]}")
    log.append(f"Testing f1 (micro, macro, weighted) = {test_stats[3]}")
    log.append("")
    log.append(f"Training auc (macro, weighted) = {train_stats[4]}")
    log.append(f"Testing auc (macro, weighted) = {test_stats[4]}")


    log.append("")
    log.append("Hyper-Parameters: \n")
    i = 0
    while i < len(params):
        log.append(f"{params[i]}{params[i+1]}")
        i += 2
    log.append("")

    # Filtering for Statistics
    
    train_x = data[0]
    train_y = data[1]
    test_x = data[2]
    test_y = data[3]
    y_trainPred = data[4]
    y_pred = data[5]

    dftrain = train_x.copy()
    dftrain["FieldSlicePrediction"] = y_trainPred #add to columns
    dftrain["FieldSliceActual"] = train_y
    dftrain = dftrain.assign(Correct = lambda x: (x["FieldSliceActual"] == x["FieldSlicePrediction"]))

    dftest = test_x.copy()
    dftest["FieldSlicePrediction"] = y_pred #add to columns
    dftest["FieldSliceActual"] = test_y
    dftest = dftest.assign(Correct = lambda x: (x["FieldSliceActual"] == x["FieldSlicePrediction"]))
    
    dfall = pd.concat([dftrain, dftest]) # add rows

    # Can either leave this code below or essentially switch it all with var=dftrain["FieldSliceActual"].value_counts()[1]
    dfTestStats = dftest.groupby(["FieldSliceActual"]).size().reset_index()
    dfTestStats = dfTestStats.rename(columns={"FieldSliceActual":"Field Slice",0:"Count of Actual"})
    dfTestStats["Count of Predicted"] = dftest.groupby(["FieldSlicePrediction"]).size().reset_index()[0]
    dftemp = dftest[dftest["Correct"] == True]
    dfTestStats["Correct"] = dftemp.groupby(["FieldSliceActual"]).size().reset_index()[0]

    dfTrainStats = dftrain.groupby(["FieldSliceActual"]).size().reset_index()
    dfTrainStats = dfTrainStats.rename(columns={"FieldSliceActual":"Field Slice",0:"Count of Actual"})
    dfTrainStats["Count of Predicted"] = dftrain.groupby(["FieldSlicePrediction"]).size().reset_index()[0]
    dftemp = dftrain[dftrain["Correct"] == True]
    dfTrainStats["Correct"] = dftemp.groupby(["FieldSliceActual"]).size().reset_index()[0]


    log.append("Accuracy Score for Predicting on Training Data: " + str('{:.4f}'.format(train_stats[0])))
    log.append("Accuracy Score for Predicting on Test Data: " + str('{:.4f}'.format(test_stats[0])))

    probs = model.predict_proba(test_x)
    colprob = colsum(probs, len(probs[0]), len(probs))
    colperc = ['{:.2f}'.format(n*100) for n in colprob]
    log.append("\nOverall Average Probabilities\n-------------------------------------" )
    log.append("Section 1: " + str(colperc[0]) + "%\nSection 2: " + str(colperc[1]) + "%\nSection 3: " + str(colperc[2]) + "%")
    log.append("Section 4: " + str(colperc[3]) + "%\nSection 5: " + str(colperc[4]) + "%")
    log.append("")

    log.append("Field Slice Counts for Training Data\n--------------------------------------------------")
    log.append("Section\tTruth\tPrediction")
    for i in range(dfTrainStats["Field Slice"].size):
        log.append(str(dfTrainStats["Field Slice"][i]) +"\t\t"+ str(dfTrainStats["Count of Actual"][i]) +"\t\t"+ str(dfTrainStats["Count of Predicted"][i]))
    log.append("Amount Correct: " + str(dftrain["Correct"].value_counts()[True]))
    log.append("Amount Incorrect: " + str(dftrain["Correct"].value_counts()[False]))    
    log.append("")

    log.append("Field Slice Counts for Testing Data\n--------------------------------------------------")
    log.append("Section\tTruth\tPrediction")
    for i in range(dfTestStats["Field Slice"].size):
        log.append(str(dfTestStats["Field Slice"][i]) +"\t\t"+ str(dfTestStats["Count of Actual"][i]) +"\t\t"+ str(dfTestStats["Count of Predicted"][i]))
    log.append("Amount Correct: " + str(dftest["Correct"].value_counts()[True]))
    log.append("Amount Incorrect: " + str(dftest["Correct"].value_counts()[False]))

    writeLog(log, modelType)



def printModel(modelType, model, train_stats, test_stats, data, params, fieldModelType):
    # weights = [w for w in model.w]
    
    # log.append("Weights:")

    # log += [",".join([str(w) for w in class_w]) for class_w in weights]

    log = list()
    print(f"Model Type: {modelType}")
    print("")
    print(f"Training Size = {len(data[0])}")
    print(f"Testing Size = {len(data[2])}")
    print("")
    print(f"Training Accuracy = {train_stats[0]}")
    print(f"Testing Accuracy = {test_stats[0]}")
    print("")
    print(f"Training Average Error = {train_stats[1]}")
    print(f"Testing Average Error = {test_stats[1]}")    
    print("")
    print(f"Training Recall = {train_stats[2]}")
    print(f"Testing Recall = {test_stats[2]}")
    print("")
    print(f"Training f1 (micro, macro, weighted) = {train_stats[3]}")
    print(f"Testing f1 (micro, macro, weighted) = {test_stats[3]}")
    print("")
    print(f"Training auc (macro, weighted) = {train_stats[4]}")
    print(f"Testing auc (macro, weighted) = {test_stats[4]}")
    
    print("")
    print("Hyper-Parameters: \n")
    i = 0
    while i < len(params):
        print(f"{params[i]}{params[i+1]}")
        i += 2
    print("")

    # Filtering for Statistics
    
    train_x = data[0]
    train_y = data[1]
    test_x = data[2]
    test_y = data[3]
    y_trainPred = data[4]
    y_pred = data[5]

    if fieldModelType in "Infield":
        s = "Slice"
    if fieldModelType in "Outfield":
        s = "Section"

    dftrain = train_x.copy()
    dftrain["Field"+s+"Prediction"] = y_trainPred #add to columns
    dftrain["Field"+s+"Actual"] = train_y
    dftrain = dftrain.assign(Correct = lambda x: (x["Field"+s+"Actual"] == x["Field"+s+"Prediction"]))

    dftest = test_x.copy()
    dftest["Field"+s+"Prediction"] = y_pred #add to columns
    dftest["Field"+s+"Actual"] = test_y
    dftest = dftest.assign(Correct = lambda x: (x["Field"+s+"Actual"] == x["Field"+s+"Prediction"]))
    
    dfall = pd.concat([dftrain, dftest]) # add rows

    # Can either leave this code below or essentially switch it all with var=dftrain["FieldSliceActual"].value_counts()[1]
    dfTestStats = dftest.groupby(["Field"+s+"Actual"]).size().reset_index()
    dfTestStats = dfTestStats.rename(columns={"Field"+s+"Actual":"Field "+s,0:"Count of Actual"})
    dfTestStats["Count of Predicted"] = dftest.groupby(["Field"+s+"Prediction"]).size().reset_index()[0]
    dftemp = dftest[dftest["Correct"] == True]
    dfTestStats["Correct"] = dftemp.groupby(["Field"+s+"Actual"]).size().reset_index()[0]

    dfTrainStats = dftrain.groupby(["Field"+s+"Actual"]).size().reset_index()
    dfTrainStats = dfTrainStats.rename(columns={"Field"+s+"Actual":"Field "+s,0:"Count of Actual"})
    dfTrainStats["Count of Predicted"] = dftrain.groupby(["Field"+s+"Prediction"]).size().reset_index()[0]
    dftemp = dftrain[dftrain["Correct"] == True]
    dfTrainStats["Correct"] = dftemp.groupby(["Field"+s+"Actual"]).size().reset_index()[0]

    dftrain = train_x.copy()
    dftrain["Field"+s+"Prediction"] = y_trainPred #add to columns
    dftrain["Field"+s+"Actual"] = train_y
    dftrain = dftrain.assign(Correct = lambda x: (x["Field"+s+"Actual"] == x["Field"+s+"Prediction"]))

    dftest = test_x.copy()
    dftest["Field"+s+"Prediction"] = y_pred #add to columns
    dftest["Field"+s+"Actual"] = test_y
    dftest = dftest.assign(Correct = lambda x: (x["Field"+s+"Actual"] == x["Field"+s+"Prediction"]))
    
    dfall = pd.concat([dftrain, dftest]) # add rows

    # Can either leave this code below or essentially switch it all with var=dftrain["FieldSliceActual"].value_counts()[1]
    dfTestStats = dftest.groupby(["Field"+s+"Actual"]).size().reset_index()
    dfTestStats = dfTestStats.rename(columns={"Field"+s+"Actual":"Field "+s,0:"Count of Actual"})
    dfTestStats["Count of Predicted"] = dftest.groupby(["Field"+s+"Prediction"]).size().reset_index()[0]
    dftemp = dftest[dftest["Correct"] == True]
    dfTestStats["Correct"] = dftemp.groupby(["Field"+s+"Actual"]).size().reset_index()[0]

    dfTrainStats = dftrain.groupby(["Field"+s+"Actual"]).size().reset_index()
    dfTrainStats = dfTrainStats.rename(columns={"Field"+s+"Actual":"Field "+s,0:"Count of Actual"})
    dfTrainStats["Count of Predicted"] = dftrain.groupby(["Field"+s+"Prediction"]).size().reset_index()[0]
    dftemp = dftrain[dftrain["Correct"] == True]
    dfTrainStats["Correct"] = dftemp.groupby(["Field"+s+"Actual"]).size().reset_index()[0]


    print("Accuracy Score for Predicting on Training Data: " + str('{:.4f}'.format(train_stats[1])))
    print("Accuracy Score for Predicting on Test Data: " + str('{:.4f}'.format(test_stats[0])))

    probs = model.predict_proba(test_x)
    colprob = colsum(probs, len(probs[0]), len(probs))
    colperc = ['{:.2f}'.format(n*100) for n in colprob]

    if fieldModelType in "Infield":
        s = "Slice"
        print("\nOverall Average Probabilities\n-------------------------------------" )
        print("Section 1: " + str(colperc[0]) + "%\nSection 2: " + str(colperc[1]) + "%\nSection 3: " + str(colperc[2]) + "%")
        print("Section 4: " + str(colperc[3]) + "%\nSection 5: " + str(colperc[4]) + "%")
        print("")

        print("Field Slice Counts for Training Data\n--------------------------------------------------")
        print("Section\tTruth\tPrediction")
        for i in range(dfTrainStats["Field Slice"].size):
            print(str(dfTrainStats["Field Slice"][i]) +"\t\t"+ str(dfTrainStats["Count of Actual"][i]) +"\t\t"+ str(dfTrainStats["Count of Predicted"][i]))
        print("Amount Correct: " + str(dftrain["Correct"].value_counts()[True]))
        print("Amount Incorrect: " + str(dftrain["Correct"].value_counts()[False]))    
        print("")

        print("Field Slice Counts for Testing Data\n--------------------------------------------------")
        print("Section\tTruth\tPrediction")
        for i in range(dfTestStats["Field Slice"].size):
            print(str(dfTestStats["Field Slice"][i]) +"\t\t"+ str(dfTestStats["Count of Actual"][i]) +"\t\t"+ str(dfTestStats["Count of Predicted"][i]))
        print("Amount Correct: " + str(dftest["Correct"].value_counts()[True]))
        print("Amount Incorrect: " + str(dftest["Correct"].value_counts()[False]))

    if fieldModelType in "Outfield":
        s = "Section"
        print("\nOverall Average Probabilities\n-------------------------------------" )
        print("Section 6: " + str(colperc[0]) + "%\nSection 7: " + str(colperc[1]) + "%\nSection 8: " + str(colperc[2]) + "%")
        print("Section 9: " + str(colperc[3]) + "%\nSection 10: " + str(colperc[4]) + "%"+ "%\nSection 11: " + str(colperc[5]) + "%")
        print("Section 12: " + str(colperc[6]) + "%\nSection 13: " + str(colperc[7]) + "%\nSection 14: " + str(colperc[8]) + "%")
        print("Section 15: " + str(colperc[9]) + "%\nSection 16: " + str(colperc[10]) + "%"+ "%\nSection 17: " + str(colperc[11]) + "%")
        print("Section 18: " + str(colperc[12]) + "%\nSection 19: " + str(colperc[13]) + "%"+ "%\nSection 20: " + str(colperc[14]) + "%")
        print("")

        print("Field Slice Counts for Training Data\n--------------------------------------------------")
        print("Section\tTruth\tPrediction")
        for i in range(dfTrainStats["Field Section"].size):
            print(str(int(dfTrainStats["Field Section"][i])+6) +"\t\t"+ str(dfTrainStats["Count of Actual"][i]) +"\t\t"+ str(dfTrainStats["Count of Predicted"][i]))
        print("Amount Correct: " + str(dftrain["Correct"].value_counts()[True]))
        print("Amount Incorrect: " + str(dftrain["Correct"].value_counts()[False]))
        print("")

        print("Field Section Counts for Testing Data\n--------------------------------------------------")
        print("Section\tTruth\tPrediction")
        for i in range(dfTestStats["Field Section"].size):
            print(str(int(dfTrainStats["Field Section"][i])+6) +"\t\t"+ str(dfTestStats["Count of Actual"][i]) +"\t\t"+ str(dfTestStats["Count of Predicted"][i]))
        print("Amount Correct: " + str(dftest["Correct"].value_counts()[True]))
        print("Amount Incorrect: " + str(dftest["Correct"].value_counts()[False]))

    
def printOutfieldModel(modelType, model, train_stats, test_stats, data, params):
    # weights = [w for w in model.w]
    
    # log.append("Weights:")

    # log += [",".join([str(w) for w in class_w]) for class_w in weights]

    log = list()
    print(f"Model Type: {modelType}")
    print("")
    print(f"Training Size = {len(data[0])}")
    print(f"Testing Size = {len(data[2])}")
    print("")
    print(f"Training Accuracy = {train_stats[0]}")
    print(f"Testing Accuracy = {test_stats[0]}")
    print("")
    print(f"Training Average Error = {train_stats[1]}")
    print(f"Testing Average Error = {test_stats[1]}")    
    print("")
    print(f"Training Recall = {train_stats[2]}")
    print(f"Testing Recall = {test_stats[2]}")
    print("")
    print(f"Training f1 (micro, macro, weighted) = {train_stats[3]}")
    print(f"Testing f1 (micro, macro, weighted) = {test_stats[3]}")
    print("")
    print(f"Training auc (macro, weighted) = {train_stats[4]}")
    print(f"Testing auc (macro, weighted) = {test_stats[4]}")
    
    print("")
    print("Hyper-Parameters: \n")
    i = 0
    while i < len(params):
        print(f"{params[i]}{params[i+1]}")
        i += 2
    print("")

    # Filtering for Statistics
    
    train_x = data[0]
    train_y = data[1]
    test_x = data[2]
    test_y = data[3]
    y_trainPred = data[4]
    y_pred = data[5]

    dftrain = train_x.copy()
    dftrain["FieldSectionPrediction"] = y_trainPred #add to columns
    dftrain["FieldSectionActual"] = train_y
    dftrain = dftrain.assign(Correct = lambda x: (x["FieldSectionActual"] == x["FieldSectionPrediction"]))

    dftest = test_x.copy()
    dftest["FieldSectionPrediction"] = y_pred #add to columns
    dftest["FieldSectionActual"] = test_y
    dftest = dftest.assign(Correct = lambda x: (x["FieldSectionActual"] == x["FieldSectionPrediction"]))
    
    dfall = pd.concat([dftrain, dftest]) # add rows

    # Can either leave this code below or essentially switch it all with var=dftrain["FieldSliceActual"].value_counts()[1]
    dfTestStats = dftest.groupby(["FieldSectionActual"]).size().reset_index()
    dfTestStats = dfTestStats.rename(columns={"FieldSectionActual":"Field Section",0:"Count of Actual"})
    dfTestStats["Count of Predicted"] = dftest.groupby(["FieldSectionPrediction"]).size().reset_index()[0]
    dftemp = dftest[dftest["Correct"] == True]
    dfTestStats["Correct"] = dftemp.groupby(["FieldSectionActual"]).size().reset_index()[0]

    dfTrainStats = dftrain.groupby(["FieldSectionActual"]).size().reset_index()
    dfTrainStats = dfTrainStats.rename(columns={"FieldSectionActual":"Field Section",0:"Count of Actual"})
    dfTrainStats["Count of Predicted"] = dftrain.groupby(["FieldSectionPrediction"]).size().reset_index()[0]
    dftemp = dftrain[dftrain["Correct"] == True]
    dfTrainStats["Correct"] = dftemp.groupby(["FieldSectionActual"]).size().reset_index()[0]


    print("Accuracy Score for Predicting on Training Data: " + str('{:.4f}'.format(train_stats[1])))
    print("Accuracy Score for Predicting on Test Data: " + str('{:.4f}'.format(test_stats[0])))

    probs = model.predict_proba(test_x)
    colprob = colsum(probs, len(probs[0]), len(probs))
    colperc = ['{:.2f}'.format(n*100) for n in colprob]
    print("\nOverall Average Probabilities\n-------------------------------------" )
    print("Section 6: " + str(colperc[0]) + "%\nSection 7: " + str(colperc[1]) + "%\nSection 8: " + str(colperc[2]) + "%")
    print("Section 9: " + str(colperc[3]) + "%\nSection 10: " + str(colperc[4]) + "%"+ "%\nSection 11: " + str(colperc[5]) + "%")
    print("Section 12: " + str(colperc[6]) + "%\nSection 13: " + str(colperc[7]) + "%\nSection 14: " + str(colperc[8]) + "%")
    print("Section 15: " + str(colperc[9]) + "%\nSection 16: " + str(colperc[10]) + "%"+ "%\nSection 17: " + str(colperc[11]) + "%")
    print("Section 18: " + str(colperc[12]) + "%\nSection 19: " + str(colperc[13]) + "%"+ "%\nSection 20: " + str(colperc[14]) + "%")
    print("")

    print("Field Slice Counts for Training Data\n--------------------------------------------------")
    print("Section\tTruth\tPrediction")
    for i in range(dfTrainStats["Field Section"].size):
        print(str(int(dfTrainStats["Field Section"][i])+6) +"\t\t"+ str(dfTrainStats["Count of Actual"][i]) +"\t\t"+ str(dfTrainStats["Count of Predicted"][i]))
    print("Amount Correct: " + str(dftrain["Correct"].value_counts()[True]))
    print("Amount Incorrect: " + str(dftrain["Correct"].value_counts()[False]))
    print("")

    print("Field Section Counts for Testing Data\n--------------------------------------------------")
    print("Section\tTruth\tPrediction")
    for i in range(dfTestStats["Field Section"].size):
        print(str(int(dfTrainStats["Field Section"][i])+6) +"\t\t"+ str(dfTestStats["Count of Actual"][i]) +"\t\t"+ str(dfTestStats["Count of Predicted"][i]))
    print("Amount Correct: " + str(dftest["Correct"].value_counts()[True]))
    print("Amount Incorrect: " + str(dftest["Correct"].value_counts()[False]))


def writeToExcelSheet(logDF, name=""):
    now = datetime.now()
    dt_string = now.strftime("_%m-%Y")
    filename = f"Logs/ModelStatistics{dt_string}.xlsx"
    exists = os.path.isfile(filename)
    hbool = False
    if not exists:
        wb = Workbook()
        ws=wb.active
        ws.title=name
        hbool = True
        #with pd.ExcelWriter(filename, mode='w') as writer:
        #    logDF.to_excel(writer, sheet_name=name)
    else:
        wb = load_workbook(filename)
        if name not in wb.sheetnames:
            hbool = True
            wb.create_sheet(name)
        ws = wb[name]

    for row in op.dataframe_to_rows(logDF, index=False, header=hbool):
        ws.append(row)
    wb.save(filename)

def writeToImageExcelSheet(picList, name=""):
    now = datetime.now()
    dt_string = now.strftime("_%m-%Y")
    filename = f"Logs/ModelStatistics{dt_string}.xlsx"
    exists = os.path.isfile(filename)
    hbool = False
    if not exists:
        wb = Workbook()
        ws=wb.active
        ws.title=name
        hbool = True
        #with pd.ExcelWriter(filename, mode='w') as writer:
        #    logDF.to_excel(writer, sheet_name=name)
    else:
        wb = load_workbook(filename)
        if name not in wb.sheetnames:
            hbool = True
            wb.create_sheet(name)
        ws = wb[name]
# Max could be 16 pics
    wspots = ['A1','G1','M1','S1','A15','G15','M15','S15','A29','G29','M29','S29','A43','G43','M43','S43']
    pspots = ['A2','G2','M2','S2','A16','G16','M16','S16','A30','G30','M30','S30','A44','G44','M44','S44']
    for pic,wspot,pspot in zip(picList,wspots[:len(picList)],pspots[:len(picList)]):
        picNameList = pic.split('_') 
        # [0] is last name, [1] is first name, [2] is Pitch Type, [3] is Batter type
        # Ex: 'Allsup_Chase_ChangeUp_LeftBatter.png'
        picNameList[3] = picNameList[3].split('.')[0]
        ws[wspot] = picNameList[2] + ' thrown to a ' + picNameList[3]
        # create an image
        img = Image("Visualization/"+pic)
        img.width /= 3.5
        img.height /= 3.5
        # add to worksheet and anchor next to cells
        ws.add_image(img, pspot)

    wb.save(filename)

def ExcelModel(modelType, model, train_stats, test_stats, data, params, fieldModelType): #need to add what kind of training splits done?
    # weights = [w for w in model.w]
    
    # log.append("Weights:")

    # log += [",".join([str(w) for w in class_w]) for class_w in weights]
    DFdic = params
    DFdic.update({'Training Size':len(data[0]), 'Testing Size':len(data[2]), 'Training Accuracy':train_stats[0], 'Testing Accuracy':test_stats[0],
                  'Training Average Error':train_stats[1], 'Testing Average Error':test_stats[1]})
    for i,num in enumerate(train_stats[2]):
        DFdic.update({("Training Recall : Section "+str(i)):num}) 
    for i,num in enumerate(test_stats[2]):
        DFdic.update({("Testing Recall : Section "+str(i)):num}) 
    DFdic.update({'Training F1(micro)':train_stats[3][0], 'Training F1(macro)':train_stats[3][1], 'Training F1(weighted)':train_stats[3][2]})
    DFdic.update({'Testing F1(micro)':test_stats[3][0], 'Testing F1(macro)':test_stats[3][1], 'Testing F1(weighted)':test_stats[3][2]})
    DFdic.update({'Training AUC(macro)':train_stats[4][0], 'Training AUC(weighted)':train_stats[4][1]})
    DFdic.update({'Testing AUC(macro)':test_stats[4][0], 'Testing AUC(weighted)':test_stats[4][1]})
    
    # Filtering for Statistics
    
    train_x = data[0]
    train_y = data[1]
    test_x = data[2]
    test_y = data[3]
    y_trainPred = data[4]
    y_pred = data[5]

    dftrain = train_x.copy()
    dftrain["FieldSlicePrediction"] = y_trainPred #add to columns
    dftrain["FieldSliceActual"] = train_y
    dftrain = dftrain.assign(Correct = lambda x: (x["FieldSliceActual"] == x["FieldSlicePrediction"]))

    dftest = test_x.copy()
    dftest["FieldSlicePrediction"] = y_pred #add to columns
    dftest["FieldSliceActual"] = test_y
    dftest = dftest.assign(Correct = lambda x: (x["FieldSliceActual"] == x["FieldSlicePrediction"]))
    dfall = pd.concat([dftrain, dftest])
    dfTestStats = dftest.groupby(["FieldSliceActual"]).size().reset_index()
    dfTestStats = dfTestStats.rename(columns={"FieldSliceActual":"Field Slice",0:"Count of Actual"})
    dfTestStats["Count of Predicted"] = dftest.groupby(["FieldSlicePrediction"]).size().reset_index()[0]
    dftemp = dftest[dftest["Correct"] == True]
    dfTestStats["Correct"] = dftemp.groupby(["FieldSliceActual"]).size().reset_index()[0]

    dfTrainStats = dftrain.groupby(["FieldSliceActual"]).size().reset_index()
    dfTrainStats = dfTrainStats.rename(columns={"FieldSliceActual":"Field Slice",0:"Count of Actual"})
    dfTrainStats["Count of Predicted"] = dftrain.groupby(["FieldSlicePrediction"]).size().reset_index()[0]
    dftemp = dftrain[dftrain["Correct"] == True]
    dfTrainStats["Correct"] = dftemp.groupby(["FieldSliceActual"]).size().reset_index()[0]
    probs = model.predict_proba(test_x)
    colprob = colsum(probs, len(probs[0]), len(probs))

    for i,num in enumerate(colprob):
        DFdic.update({("Section "+str(i)+" Probability"):num}) 

    df = pd.DataFrame(data=DFdic, index=[0])
    writeToExcelSheet(df,modelType)


def colsum(arr, n, m):
    coll = [0] * len(arr)
    for i in range(n):
        su = 0;
        for j in range(m):
            su += arr[j][i]
        coll[i] = su/m
    return coll 

def excelAverages(modelType, sColumns, sColumnsLetter):
    # TODO
    # This is meant to take all the values from the 30 runs and average them and output them to another sheet of averages for different models
    # Then will need to do this for all the models
    # Can take this and put it into an excelAverages function

    now = datetime.now()
    dt_string = now.strftime("_%m-%Y")
    filename = f"Logs/ModelStatistics{dt_string}.xlsx"
    wb = openpyxl.load_workbook(filename)
    #first_sheet = wb.get_sheet_names()[0]
    worksheet = wb.get_sheet_by_name(modelType)
    # These are the columns from the excel sheet that we want to average and put on Averages page
    

    # here you iterate over the rows in the specific column
    # could also add the averages of just the certain permutation of hyperparams to see how its doing
    # So an overall average and then
    avgList = [0] * len(sColumns)
    for row in range(2,worksheet.max_row+1): 
        for i, column in enumerate(sColumnsLetter):  #Here you can add or reduce the columns
            cell_name = "{}{}".format(column, row)
            if str(type(worksheet[cell_name].value)) not in 'str':
                #print(str(i) + ":" + str(worksheet[cell_name].value))
                avgList[i] += worksheet[cell_name].value # the value of the specific cell

    avgListEnd = [x/(worksheet.max_row-1) for x in avgList]
    avgListEnd.insert(0, modelType)
    sColumns.insert(0, 'Model Type')
    # Still needs to export to the workbooks like everything else
    if (config['LOGGING']['Debug'] == 'True'):
        print("printing statistics...")
        print(avgListEnd)
    if (config['LOGGING']['Excel'] == 'True'):
        print("exporting statistics to Excel...")
        df = pd.DataFrame([avgListEnd], columns=sColumns)
        writeToExcelSheet(df,"Average Statistics")