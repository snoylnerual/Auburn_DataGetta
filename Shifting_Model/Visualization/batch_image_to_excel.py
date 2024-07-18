import os
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment  # Import Font and Alignment

# Define a function to update the cell style for headers
def style_cell(sheet, cell, text, bold=True, size=14):
    sheet[cell] = text
    sheet[cell].font = Font(bold=bold, size=size)
    sheet[cell].alignment = Alignment(horizontal='center',vertical='center')
    # Expand the cell width and height based on text size, if necessary
    # This is a simplistic approach; you might need to adjust it.
    width = len(text) + 2  # Basic calculation, adjust as needed
    height = 20  # Adjust as needed
    sheet.column_dimensions[(cell[0])].width = width
    sheet.row_dimensions[int(cell[1:])].height = height

def insert_image(sheet, image_path, cell):
    img = Image(image_path)
    # Scale the image if needed, here set as per your needs
    img.width, img.height = 1094, 820  # You might need to adjust these values
    img.width = img.width / 2
    img.height = img.height / 2
    # Insert the image into the sheet
    sheet.add_image(img, cell)

    # Adjust column width and row height to fit the image
    column_letter = (cell[0])
    sheet.column_dimensions[column_letter].width = img.width / 7  # Adjusting column width
    sheet.row_dimensions[int(cell[1:])].height = img.height / 1.2  # Adjusting row height


# Function to process each image file
def process_image(workbook, filename, folder_path):
    parts = filename.split('_')
    if len(parts) != 4:
        print(f"Invalid filename format: {filename}")
        return

    lastname, firstname, pitch, hand = parts
    hand = hand.split('.')[0]  # Remove the file extension

    # Determine the worksheet name
    sheet_name = f"{lastname}_{firstname}"
    sheet = ''
    # Create or get the worksheet
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        # player_name = f"{firstname} {lastname}".title()
        # style_cell(sheet, 'A1', player_name, bold=True, size=24)
        # sheet.column_dimensions['A'].width = 40  # Adjusting column width
        # sheet.row_dimensions[1].height = 30 # Adjusting row height
        # sheet.column_dimensions['C'].width = 40  # Adjusting column width
    else:
        sheet = workbook.create_sheet(sheet_name)
        # Set the pitch headers and "Left" and "Right" labels
        pitches = ['FASTBALL', 'SINKER', 'CHANGEUP', #'FOURSEAMFASTBALL' == FASTBALL, 'TWOSEAMFASTBALL' == SINKER
                   'SLIDER', 'CURVEBALL', 'CUTTER', 'SPLITTER']
        style_cell(sheet, 'C1', 'Batter Hand', bold=True, size=14)
        style_cell(sheet, 'B2', 'Left', bold=True, size=14)
        style_cell(sheet, 'D2', 'Right', bold=True, size=14)
        player_name = f"{firstname} {lastname}".title()
        style_cell(sheet, 'A1', player_name, bold=True, size=24)
        sheet.column_dimensions['A'].width = 40  # Adjusting column width
        sheet.row_dimensions[1].height = 30 # Adjusting row height

    # Insert the image into the correct cell
    cell = "" 
    if pitch.upper() == 'FASTBALL':
        cell = 'B3' if hand.upper() == 'LEFTBATTER' else 'D3'
        style_cell(sheet, 'C3', 'Fastball', bold=True, size=14)
    elif pitch.upper() == 'FOURSEAMFASTBALL':
        cell = 'B4' if hand.upper() == 'LEFTBATTER' else 'D4'
        style_cell(sheet, 'C4', 'FourSeamFastball', bold=True, size=14)
    elif pitch.upper() == 'TWOSEAMFASTBALL':
        cell = 'B5' if hand.upper() == 'LEFTBATTER' else 'D5'
        style_cell(sheet, 'C5', 'TwoSeamFastBall', bold=True, size=14)
    elif pitch.upper() == 'SINKER':
        cell = 'B6' if hand.upper() == 'LEFTBATTER' else 'D6'
        style_cell(sheet, 'C6', 'Sinker', bold=True, size=14)
    elif pitch.upper() == 'CHANGEUP':
        cell = 'B7' if hand.upper() == 'LEFTBATTER' else 'D7'
        style_cell(sheet, 'C7', 'Changeup', bold=True, size=14)
    elif pitch.upper() == 'SLIDER':
        cell = 'B8' if hand.upper() == 'LEFTBATTER' else 'D8'
        style_cell(sheet, 'C8', 'Slider', bold=True, size=14)
    elif pitch.upper() == 'CURVEBALL':
        cell = 'B9' if hand.upper() == 'LEFTBATTER' else 'D9'
        style_cell(sheet, 'C9', 'Curveball', bold=True, size=14)
    elif pitch.upper() == 'CUTTER':
        cell = 'B10' if hand.upper() == 'LEFTBATTER' else 'D10'
        style_cell(sheet, 'C10', 'Cutter', bold=True, size=14)
    elif pitch.upper() == 'SPLITTER':
        cell = 'B11' if hand.upper() == 'LEFTBATTER' else 'D11'
        style_cell(sheet, 'C11', 'Splitter', bold=True, size=14)
    else:
        print(pitch.upper())


    if cell != "":
        insert_image(sheet, os.path.join(folder_path, filename), cell)
        sheet.column_dimensions['C'].width = 25  # Adjusting column width



# Main program starts here
def create_excel():
    # Path to the target directory
    folder_path = os.path.abspath(os.path.join(os.curdir, 'Output'))
    print("Looking in:", folder_path)

    # folder_path = 'C:\Users\Trent\Desktop\Senior Design\shifting_model\Output'  # Change this to the path of your images folder
    excel_path = 'Pitcher_Averages.xlsx'  # Change this to your Excel file path, it can be new or existing

    # Load or create the workbook
    if os.path.exists(excel_path):
        os.remove(excel_path)
        wb = Workbook()
    else:
        wb = Workbook()

    # Process each image in the folder
    for filename in os.listdir(folder_path):
        if filename.endswith(('.png', '.jpg', '.jpeg')):  # Add more file types if needed
            process_image(wb, filename, folder_path)

    # Save the workbook
    default_sheet_name = wb.sheetnames[0]  # Get the name of the first sheet
    if default_sheet_name == 'Sheet':  # Change 'Sheet' to your default sheet name if different
        default_sheet = wb[default_sheet_name]
        wb.remove(default_sheet)

    wb.save(excel_path)
    print(f"Workbook saved at {excel_path}")
